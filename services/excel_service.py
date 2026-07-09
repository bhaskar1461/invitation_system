import os
import re
from openpyxl import load_workbook
from flask import current_app

from models import db
from models.guest import Guest
from services.guest_service import GuestService

class ExcelService:
    @staticmethod
    def process_import(file_path):
        """
        Parses an Excel spreadsheet, validates columns and data format,
        extracts optional rollno and mobile, generates QR codes, and inserts into event_qr_codes.
        """
        summary = {
            'success': False,
            'imported_count': 0,
            'skipped_duplicates': 0,
            'invalid_emails': 0,
            'errors': [],
            'skipped_details': []
        }
        
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            sheet = wb.active
            
            # Read header row
            rows_iter = sheet.iter_rows(values_only=True)
            try:
                headers = next(rows_iter)
            except StopIteration:
                summary['errors'].append("The Excel sheet is empty.")
                return summary

            # Normalize headers
            if not headers:
                summary['errors'].append("Could not read headers.")
                return summary
                
            headers_clean = [str(h).strip().lower() for h in headers if h is not None]
            
            # Check for required headers
            if len(headers_clean) < 2 or 'name' not in headers_clean or 'email' not in headers_clean:
                summary['errors'].append("Invalid headers. Excel file must contain 'Name' and 'Email' columns.")
                return summary
                
            name_idx = headers_clean.index('name')
            email_idx = headers_clean.index('email')
            
            # Dynamically look up optional rollno and mobile columns
            rollno_idx = None
            for idx, h in enumerate(headers_clean):
                clean_h = re.sub(r'[^a-z0-9]', '', h)
                if any(kw in clean_h for kw in ('roll', 'id', 'regno', 'admissionno')):
                    rollno_idx = idx
                    break

            mobile_idx = None
            for idx, h in enumerate(headers_clean):
                clean_h = re.sub(r'[^a-z0-9]', '', h)
                if any(kw in clean_h for kw in ('mobile', 'phone', 'contact', 'cell', 'phno', 'tel')):
                    mobile_idx = idx
                    break
            
            existing_guests = {g.email.lower(): g for g in Guest.query.filter(Guest.email.is_not(None)).all()}
            existing_codes = {g.qr_code for g in existing_guests.values()}
            
            seen_emails = set()
            row_num = 1
            guests_to_add = []
            guests_to_update = []
            
            def get_unique_code_in_memory():
                import random
                for _ in range(10):
                    code = str(random.randint(100000, 999999))
                    if code not in existing_codes:
                        existing_codes.add(code)
                        return code
                raise RuntimeError("Failed to generate a unique 6-digit code.")
            
            for row in rows_iter:
                row_num += 1
                
                # Check if row is empty
                if not row or len(row) <= max(name_idx, email_idx):
                    continue
                    
                name_val = row[name_idx]
                email_val = row[email_idx]
                
                name = str(name_val).strip() if name_val is not None else ""
                email = str(email_val).strip().lower() if email_val is not None else ""
                
                # Skip fully empty rows
                if not name and not email:
                    continue
                
                # Validate name and email are present
                if not name or not email:
                    summary['invalid_emails'] += 1
                    summary['skipped_details'].append({
                        'row': row_num,
                        'name': name or "[Missing]",
                        'email': email or "[Missing]",
                        'reason': "Name or Email is missing"
                    })
                    continue
                
                # Validate email format
                if not re.match(r"[^@]+@[^@]+\.[^@]+", email):
                    summary['invalid_emails'] += 1
                    summary['skipped_details'].append({
                        'row': row_num,
                        'name': name,
                        'email': email,
                        'reason': "Invalid email format"
                    })
                    continue
                
                # Check for duplicates within this file
                if email in seen_emails:
                    summary['skipped_duplicates'] += 1
                    summary['skipped_details'].append({
                        'row': row_num,
                        'name': name,
                        'email': email,
                        'reason': "Duplicate email in spreadsheet"
                    })
                    continue
                
                # Extract optional fields
                rollno = None
                if rollno_idx is not None and len(row) > rollno_idx:
                    rollno_val = row[rollno_idx]
                    rollno = str(rollno_val).strip() if rollno_val is not None else ""

                mobile = None
                if mobile_idx is not None and len(row) > mobile_idx:
                    mobile_val = row[mobile_idx]
                    mobile = str(mobile_val).strip() if mobile_val is not None else ""
                
                # Check if guest already exists in database
                if email in existing_guests:
                    existing_guest = existing_guests[email]
                    updated = False
                    if mobile and existing_guest.mobile != mobile:
                        existing_guest.mobile = mobile
                        updated = True
                    if rollno and existing_guest.rollno != rollno:
                        existing_guest.rollno = rollno
                        updated = True
                    if updated:
                        db.session.add(existing_guest)
                        guests_to_update.append(existing_guest)
                    summary['imported_count'] += 1
                    seen_emails.add(email)
                    continue
                
                # Add to file set to prevent duplicates inside sheet
                seen_emails.add(email)
                
                # Create Guest and append to transaction
                try:
                    code = get_unique_code_in_memory()
                    new_guest = Guest(
                        guest_name=name.strip(),
                        rollno=rollno or "",
                        mobile=mobile or "",
                        email=email,
                        qr_code=code,
                        qr_image=None,
                        invite_sent=False,
                        status='ACTIVE'
                    )
                    db.session.add(new_guest)
                    guests_to_add.append(new_guest)
                    summary['imported_count'] += 1
                except Exception as e:
                    current_app.logger.error(f"Error creating guest on row {row_num}: {str(e)}")
                    summary['skipped_details'].append({
                        'row': row_num,
                        'name': name,
                        'email': email,
                        'reason': f"Database error: {str(e)}"
                    })
            
            wb.close()
            
            # Commit all inserts and updates in a single transaction
            if guests_to_add or guests_to_update:
                try:
                    db.session.commit()
                except Exception as commit_err:
                    db.session.rollback()
                    current_app.logger.error(f"Failed to commit bulk import: {str(commit_err)}")
                    summary['errors'].append(f"Database error: failed to save guest list. {str(commit_err)}")
                    summary['imported_count'] = 0
                    summary['success'] = False
                    return summary
                
                # After successful commit, trigger bulk emails ONLY for NEW guests
                if guests_to_add:
                    try:
                        from services.email_service import EmailService
                        guest_ids = [g.id for g in guests_to_add]
                        EmailService.send_bulk_invitations(guest_ids)
                    except Exception as queue_err:
                        current_app.logger.error(f"Failed to queue bulk emails during import: {str(queue_err)}")
                        summary['errors'].append(f"Imported successfully, but background email dispatch failed to queue: {str(queue_err)}")
                    
            summary['success'] = True
            
        except Exception as e:
            current_app.logger.error(f"Excel reading failed: {str(e)}")
            summary['errors'].append(f"Failed to read spreadsheet file: {str(e)}")
            
        return summary
